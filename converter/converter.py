import re
import openpyxl

def m3u_to_excel(m3u_file, excel_file):
    m3u_file = m3u_file if m3u_file.endswith(".m3u") else m3u_file + ".m3u"
    excel_file = excel_file if excel_file.endswith(".xlsx") else excel_file + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IPTV Channels"
    ws.append([
        "Channel Name", "tvg-name", "tvg-id", "Logo",
        "tvg-chno", "Group", "http-referrer", "http-user-agent", "URL"
    ])

    with open(m3u_file, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    channel_name = tvg_name = tvg_id = logo = chno = group = referrer = ua = url = ""

    for i, line in enumerate(lines):
        line = line.strip()

        # Proses hanya baris EXTINF
        if line.startswith("#EXTINF"):
            tvg_name = re.search(r'tvg-name="([^"]*)"', line)
            tvg_id = re.search(r'tvg-id="([^"]*)"', line)
            logo = re.search(r'tvg-logo="([^"]*)"', line)
            chno = re.search(r'tvg-chno="([^"]*)"', line)
            group = re.search(r'group-title="([^"]*)"', line)
            referrer = re.search(r'http-referrer="([^"]*)"', line)
            ua = re.search(r'http-user-agent="([^"]*)"', line)

            tvg_name = tvg_name.group(1) if tvg_name else ""
            tvg_id = tvg_id.group(1) if tvg_id else ""
            logo = logo.group(1) if logo else ""
            chno = chno.group(1) if chno else ""
            group = group.group(1) if group else ""
            referrer = referrer.group(1) if referrer else ""
            ua = ua.group(1) if ua else ""

            # Ambil nama channel setelah koma terakhir
            if "," in line:
                channel_name = line.split(",", 1)[-1].strip()

            # URL ada di baris berikutnya
            if i + 1 < len(lines):
                url = lines[i + 1].strip()
                if url and not url.startswith("#"):
                    ws.append([channel_name, tvg_name, tvg_id, logo, chno, group, referrer, ua, url])
                    channel_name = tvg_name = tvg_id = logo = chno = group = referrer = ua = url = ""

    ws.freeze_panes = "A2"
    wb.save(excel_file)
    print(f"✅ Data berhasil diexport ke {excel_file}")


def excel_to_m3u(excel_file, m3u_file):
    excel_file = excel_file if excel_file.endswith(".xlsx") else excel_file + ".xlsx"
    m3u_file = m3u_file if m3u_file.endswith(".m3u") else m3u_file + ".m3u"

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    with open(m3u_file, "w", encoding="utf-8") as f:
        f.write("#EXTM3U\n")  # tetap ditulis di awal agar valid

        for row in ws.iter_rows(min_row=2, values_only=True):
            (
                channel_name, tvg_name, tvg_id, logo,
                chno, group, referrer, ua, url
            ) = row

            extinf = (
                f'#EXTINF:-1 tvg-name="{tvg_name or ""}" '
                f'tvg-id="{tvg_id or ""}" '
                f'tvg-logo="{logo or ""}" '
                f'tvg-chno="{chno or ""}" '
                f'group-title="{group or ""}" '
                f'http-referrer="{referrer or ""}" '
                f'http-user-agent="{ua or ""}",{channel_name or ""}'
            )

            f.write(extinf + "\n")
            f.write((url or "") + "\n")

    print(f"✅ Playlist berhasil dibuat: {m3u_file}")


def main():
    print("=== Converter M3U <-> Excel ===")
    print("1. Convert M3U to Excel")
    print("2. Convert Excel to M3U")
    pilihan = input("Pilih (1/2): ").strip()

    if pilihan == "1":
        m3u_file = input("Masukkan nama file M3U (tanpa .m3u): ").strip()
        excel_file = input("Masukkan nama file output Excel (tanpa .xlsx): ").strip()
        m3u_to_excel(m3u_file, excel_file)
    elif pilihan == "2":
        excel_file = input("Masukkan nama file Excel (tanpa .xlsx): ").strip()
        m3u_file = input("Masukkan nama file output M3U (tanpa .m3u): ").strip()
        excel_to_m3u(excel_file, m3u_file)
    else:
        print("❌ Pilihan tidak valid")


if __name__ == "__main__":
    main()
